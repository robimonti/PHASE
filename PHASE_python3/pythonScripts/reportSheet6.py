import sys
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, numbers

def format_excel_sheet6(report_path, sheet_name, figure_paths):
    # Load the workbook and select the sheet
    wb = openpyxl.load_workbook(report_path)
    sheet = wb[sheet_name]

    # Convert figure paths string into a list
    figure_paths = figure_paths.split(',')

    # Initialize starting row
    start_row = 9  # First figure position

    # General formatting for the table
    subtitle_row = 5  # Start of the table (below figure space)
    for row in sheet.iter_rows(min_row=subtitle_row+1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            if cell.value:
                cell.font = Font(size=12)
                cell.alignment = Alignment(horizontal='left', vertical='center')

    # Insert the figures dynamically
    for i, fig_path in enumerate(figure_paths):
        # Insert the figure
        figure_position = f"A{start_row}"  # Adjust dynamically
        img = Image(fig_path)
        img.width = 900  # Set figure width
        img.height = 450  # Set figure height
        sheet.add_image(img, figure_position)

        # Apply number format for displacement time series (columns after metadata)
        data_start_row = subtitle_row + 2  # Data starts two rows after the subtitle
        for row in sheet.iter_rows(min_row=data_start_row, max_row=data_start_row+1, min_col=2, max_col=sheet.max_column):
            for cell in row:
                if cell.value:
                    cell.font = Font(size=12)
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                    # Apply number format for displacement time series
                    cell.number_format = '0.000'  # Three decimals

        # Format ID
        sheet[f"A{data_start_row}"].font = Font(size=13, bold=True)

        # Move to the next block (25 rows below for figures, plus extra space for tables)
        start_row += 27
        subtitle_row += 27


    # Set the width of columns
    sheet.column_dimensions['A'].width = 9
    sheet.column_dimensions['D'].width = 15
    sheet.column_dimensions['E'].width = 15

    # Format the title (A1)
    sheet['A1'].font = Font(size=20, bold=True)
    sheet['A1'].alignment = Alignment(horizontal='left', vertical='center')
    # Format the first subtitle (A2)
    sheet['A2'].font = Font(size=16, bold=False)
    sheet['A2'].alignment = Alignment(horizontal='left', vertical='center')
    # Format the second subtitle (A3)
    sheet['A3'].font = Font(size=16, bold=False)
    sheet['A3'].alignment = Alignment(horizontal='left', vertical='center')
    # Format the third subtitle (A5)
    sheet['A5'].font = Font(size=18, bold=True)
    sheet['A5'].alignment = Alignment(horizontal='left', vertical='center')

    # Save the workbook
    wb.save(report_path)
    print(f"Formatted sheet '{sheet_name}' and updated Excel file: {report_path}")

if __name__ == "__main__":
    # Parse arguments from MATLAB
    report_file = sys.argv[1]
    sheet_name = sys.argv[2]
    figure_paths = sys.argv[3]

    # Format the second sheet
    format_excel_sheet6(report_file, sheet_name, figure_paths)
