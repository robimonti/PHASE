import sys
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, numbers, PatternFill

def format_excel_sheet5(report_path, sheet_name):
    # Load the workbook and select the sheet
    wb = openpyxl.load_workbook(report_path)
    sheet = wb[sheet_name]

    # General formatting for the table
    subtitle_row = 5  # Start of the table (below figure space)
    for row in sheet.iter_rows(min_row=subtitle_row+1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            if cell.value is not None:
                cell.font = Font(size=12)
                cell.alignment = Alignment(horizontal='left', vertical='center')
    # Info
    for row in sheet.iter_rows(min_row=subtitle_row+2, max_row=sheet.max_row, min_col=2, max_col=5):
        for cell in row:
            if cell.value is not None:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.number_format = '0.000'  # Three decimals
    # Risks
    for row in sheet.iter_rows(min_row=subtitle_row+2, max_row=sheet.max_row, min_col=6, max_col=sheet.max_column):
        for cell in row:
            if cell.value is not None:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.number_format = '0.0'  # One decimals

    # Color fills based on global risk values
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    purple_fill = PatternFill(start_color="800080", end_color="800080", fill_type="solid")

    # Global risk is in column 10 (index 10)
    for row in sheet.iter_rows(min_row=subtitle_row+2, max_row=sheet.max_row, min_col=10, max_col=10):
        for cell in row:
            risk_value = cell.value
            cell.font = Font(size=14, bold=True)
            if risk_value < 50:
                cell.fill = green_fill
            elif 50 <= risk_value < 100:
                cell.fill = yellow_fill
            elif 100 <= risk_value < 125:
                cell.fill = orange_fill
            elif risk_value >= 125:
                cell.fill = purple_fill

    # Format the title (A1)
    sheet['A1'].font = Font(size=20, bold=True)
    sheet['A1'].alignment = Alignment(horizontal='left', vertical='center')
    # Format the first subtitle (A2)
    sheet['A2'].font = Font(size=16, bold=False)
    sheet['A2'].alignment = Alignment(horizontal='left', vertical='center')
    # Format the second subtitle (A3)
    sheet['A3'].font = Font(size=16, bold=False)
    sheet['A3'].alignment = Alignment(horizontal='left', vertical='center')

    # Format the fourth subtitle (A5)
    sheet['A5'].font = Font(size=18, bold=True)
    sheet['A5'].alignment = Alignment(horizontal='left', vertical='center')

    # Save the workbook
    wb.save(report_path)
    print(f"Formatted sheet '{sheet_name}' and updated Excel file: {report_path}")

if __name__ == "__main__":
    # Parse arguments from MATLAB
    report_file = sys.argv[1]
    sheet_name = sys.argv[2]

    # Format the second sheet
    format_excel_sheet5(report_file, sheet_name)
