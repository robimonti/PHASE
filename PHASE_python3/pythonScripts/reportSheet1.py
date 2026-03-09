import sys
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment

def format_excel(report_path, figures_and_positions, sheet_name):
    # Load the workbook and select the specified sheet
    wb = openpyxl.load_workbook(report_path)
    sheet = wb[sheet_name]

    # Parse figures and positions (input is in "path1,pos1,width1,height1;path2,pos2,width2,height2;...")
    for fig_pos in figures_and_positions.split(';'):
        figure_path, position, width, height = fig_pos.split(',')

        # Insert the image
        img = Image(figure_path)
        img.width = int(width)  # Set width in pixels
        img.height = int(height)  # Set height in pixels
        sheet.add_image(img, position)

    # General text formatting for other cells
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=2):
        for cell in row:
            if cell.value:
                cell.font = Font(size=12)
                cell.alignment = Alignment(horizontal='left', vertical='center')

    # Apply formatting
    # Title (Cell A1)
    sheet['A1'].font = Font(size=20, bold=True)
    sheet['A1'].alignment = Alignment(horizontal='left', vertical='center')
    # Subtitle (Cell A2)
    sheet['A2'].font = Font(size=14, bold=True)
    sheet['A2'].alignment = Alignment(horizontal='left', vertical='center')
    # Datetime (Cell A3)
    sheet['A3'].font = Font(size=14, bold=False)
    sheet['A3'].alignment = Alignment(horizontal='left', vertical='center')

    # Municipality (Cell A5)
    sheet['A5'].font = Font(size=16, bold=True)
    sheet['A5'].alignment = Alignment(horizontal='left', vertical='center')
    # Country (Cell A6)
    sheet['A6'].font = Font(size=16, bold=True)
    sheet['A6'].alignment = Alignment(horizontal='left', vertical='center')

    # Modelling (Cell A8)
    sheet['A8'].font = Font(size=14, bold=True)
    sheet['A8'].alignment = Alignment(horizontal='left', vertical='center')

    # Other headers
    headers = ['A10', 'A15', 'A19', 'A23']  # Example header rows
    for header in headers:
        sheet[header].font = Font(size=13, bold=True)
        sheet[header].alignment = Alignment(horizontal='left', vertical='center')

    # Save the updated workbook
    wb.save(report_path)
    print(f"Updated Excel file saved at: {report_path}")

if __name__ == "__main__":
    # Parse arguments from MATLAB
    report_file = sys.argv[1]
    figures_and_positions = sys.argv[2]  # e.g., "path1,pos1,width1,height1;path2,pos2,width2,height2"
    sheet_name = sys.argv[3]

    # Call the formatting function
    format_excel(report_file, figures_and_positions, sheet_name)
