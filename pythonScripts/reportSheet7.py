import sys
import openpyxl
from openpyxl.styles import Font, Alignment

def format_excel_sheet7(report_path, sheet_name):
    # Load the workbook and select the sheet
    wb = openpyxl.load_workbook(report_path)
    sheet = wb[sheet_name]

    # Format the title (A1)
    if sheet['A1'].value:
        sheet['A1'].font = Font(size=20, bold=True)
        sheet['A1'].alignment = Alignment(horizontal='left', vertical='center')
        
    # Format the first subtitle (A2)
    if sheet['A2'].value:
        sheet['A2'].font = Font(size=16, bold=False)
        sheet['A2'].alignment = Alignment(horizontal='left', vertical='center')

    # The table headers start at row 4 (since the placeholder array has 3 rows)
    header_row = 4
    for cell in sheet[header_row]:
        if cell.value:
            cell.font = Font(size=13, bold=True)
            cell.alignment = Alignment(horizontal='left', vertical='center')

    # General formatting for the table data (Row 5 onwards)
    for row in sheet.iter_rows(min_row=header_row+1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            if cell.value is not None:
                cell.font = Font(size=12)
                cell.alignment = Alignment(horizontal='left', vertical='center')
                
                # Apply number formats based on the column index
                if isinstance(cell.value, (int, float)):
                    if cell.column >= 6: 
                        # Displacement data (Column F onwards) -> 3 decimals
                        cell.number_format = '0.000'
                    elif cell.column in [2, 3]: 
                        # Lat / Lon -> 6 decimals for coordinate precision
                        cell.number_format = '0.000000'
                    elif cell.column in [4, 5]: 
                        # UTM X / Y -> 2 decimals
                        cell.number_format = '0.00'

    # Set the width of columns for cleaner readability
    sheet.column_dimensions['A'].width = 10  # ID
    sheet.column_dimensions['B'].width = 15  # Lat
    sheet.column_dimensions['C'].width = 15  # Lon
    sheet.column_dimensions['D'].width = 15  # X (UTM)
    sheet.column_dimensions['E'].width = 15  # Y (UTM)

    # Save the workbook
    wb.save(report_path)
    print(f"Formatted sheet '{sheet_name}' and updated Excel file: {report_path}")

if __name__ == "__main__":
    # Parse arguments from MATLAB (No figure paths needed)
    report_file = sys.argv[1]
    sheet_name = sys.argv[2]

    # Format the target sheet
    format_excel_sheet7(report_file, sheet_name)